import json
import os
import re
import math
from typing import Dict, List, Any, Optional
from models.task_models import WorkbookData, SheetData, CellData
from utils.calc_utils import clean_num, get_column_index, poisson_helper, normalize_key
from openpyxl.utils import get_column_letter

def load_scenarios():
    config_path = os.path.join(os.path.dirname(__file__), '..', 'vlookup_scenarios.json')
    with open(config_path, 'r') as f:
        return json.load(f).get("scenarios", {})

class TaskAnalyzer:
    def __init__(self, files_data, scenario):
        self.files_data = files_data
        self.scenario = scenario
        self.report = {"goal": scenario.get("description"), "diagnostics": [], "plan": []}

    def _add_diagnostic(self, severity, message, code):
        self.report["diagnostics"].append({"severity": severity, "message": message, "code": code})

    def analyze(self):
        detailed_plan = []
        for col in self.scenario.get("target_columns", []):
            label = col.get("step_label") or f"Processing {col.get('header')}"
            if "ACRD" in self.scenario.get("description", "") and "Calculate based on ACRD" in label:
                label = "Calculate Recommended Qty"
            detailed_plan.append({
                "phase": col.get("phase", "Modeling"),
                "goal": label,
                "agent": col.get("agent", "Data Mapping Agent"),
                "tool": col.get("tool", "CrossRef_Engine"),
                "status": "pending"
            })
        self.report["plan"] = detailed_plan

        source_file_key = self.scenario.get("source_file")
        source_data = self.files_data.get(source_file_key)
        if not source_data:
            self._add_diagnostic("CRITICAL", f"File '{source_file_key}' missing.", "file_missing")
            return self.report
            
        return self.report

def match_scenario(scenarios, prompt):
    prompt_lower = prompt.lower()
    for name, s in scenarios.items():
        if any(kw in prompt_lower for kw in s.get("keywords", [])):
            return s
    return scenarios.get("VLOOKUP")

def generate_vlookup_workbooks(files_data: Dict[str, List[List[Any]]], prompt: str, auto_fix: Optional[Dict] = None, command: str = None, separator: str = ";"):
    # Apply auto_fix (Resolve feature)
    param_key = next((k for k in files_data.keys() if k.lower() == 'parameters'), None)
    if auto_fix and param_key:
        print(f"DEBUG: Applying auto_fix to {param_key}: {auto_fix}")
        try:
            cell, val = auto_fix.get("cell", ""), auto_fix.get("value")
            if cell and val is not None:
                match = re.match(r"([A-Z]+)(\d+)", cell)
                if match:
                    col_str, row_str = match.groups()
                    c_idx = 0
                    for char in col_str: c_idx = c_idx * 26 + (ord(char) - ord('A'))
                    old_val = files_data[param_key][int(row_str) - 1][c_idx]
                    files_data[param_key][int(row_str) - 1][c_idx] = val
                    print(f"DEBUG: Updated cell {cell} from {old_val} to {val}")
        except Exception as e: print(f"DEBUG: Auto-fix error: {e}")

    scenarios = load_scenarios()
    scenario = scenarios.get(command)
    if not scenario: scenario = match_scenario(scenarios, prompt)
    if not scenario: return None, None, None

    normalized_files = {normalize_key(k): v for k, v in files_data.items()}
    for k, v in normalized_files.items():
        if v: print(f"DEBUG FILE STRUCTURE: '{k}' headers: {v[0][:10]} | Sample Row 1: {v[1][:10] if len(v)>1 else 'N/A'}")

    source_rows = files_data[scenario["source_file"]]
    source_headers = [str(h).strip().upper() for h in source_rows[0]]

    # 1. Pre-build lookup dictionaries
    for col in scenario["target_columns"]:
        if col.get("type") == "vlookup":
            l_file = normalize_key(col.get("lookup_file", ""))
            l_data = normalized_files.get(l_file)
            d = {}
            if l_data:
                mr = col.get("manual_range")
                if mr:
                    s_row_idx, e_row_idx = mr.get("start_row", 1), mr.get("end_row", 99999)
                    for r_i in range(s_row_idx-1, min(e_row_idx, len(l_data))):
                        r = l_data[r_i]
                        if len(r) > max(mr["key_col_idx"], mr["val_col_idx"]):
                            k = str(r[mr["key_col_idx"]]).strip()
                            if k.endswith(".0"): k = k[:-2]
                            d[k] = r[mr["val_col_idx"]]
                else:
                    l_headers = [str(h).strip().upper() for h in l_data[0]]
                    k_idx = col.get("lookup_key_col_idx", 0)
                    v_idx = col.get("target_col_idx")
                    if v_idx is None and "val_col" in col:
                        v_idx = get_column_index(l_headers, col["val_col"].upper())
                    if v_idx is None: v_idx = 1
                    for r in l_data[1:]:
                        if len(r) > max(k_idx, v_idx):
                            k = str(r[k_idx]).strip()
                            if k.endswith(".0"): k = k[:-2]
                            d[k] = r[v_idx]
            lookup_info.append({"dict": d})
        elif col.get("type") == "mock_file_list":
            f_path = col.get("file_path")
            lines = []
            if f_path:
                full_path = os.path.join(os.path.dirname(__file__), '..', f_path)
                if os.path.exists(full_path):
                    with open(full_path, 'r') as f:
                        lines = [l.strip() for l in f.readlines()]
            lookup_info.append({"lines": lines})
        else: lookup_info.append({})

    # 2. Cache for ACRD / Inhouse
    pool_qty_map = {}
    pool_key = next((k for k in normalized_files.keys() if 'pool' in k), 'inhousepoolinfo')
    p_data = normalized_files.get(pool_key)
    if p_data:
        for r in p_data:
            if len(r) > 3 and r[0] and str(r[0]).strip().upper() not in ["P/N", "PN", "PART NUMBER", "REQUESTED PN"]:
                pn_candidate = str(r[0]).strip().upper()
                if len(pn_candidate) < 5 or " " in pn_candidate: continue 
                if pn_candidate.endswith(".0"): pn_candidate = pn_candidate[:-2]
                try: pool_qty_map[pn_candidate] = clean_num(r[3])
                except: pass

    acrd_map = {}
    acrd_key = next((k for k in normalized_files.keys() if 'acrd' in k), 'acrd')
    a_data = normalized_files.get(acrd_key)
    if a_data:
        for r in a_data:
            if len(r) >= 3 and r[0] and str(r[0]).strip().upper() not in ["P/N", "PN", "PART NUMBER", "REQUESTED PN"]:
                req_pn = str(r[0]).strip().upper()
                rep_pn = str(r[2]).strip().upper()
                if len(req_pn) < 5 or " " in req_pn: continue
                if req_pn.endswith(".0"): req_pn = req_pn[:-2]
                if rep_pn.endswith(".0"): rep_pn = rep_pn[:-2]
                if req_pn not in acrd_map: acrd_map[req_pn] = []
                if rep_pn not in acrd_map[req_pn]: acrd_map[req_pn].append(rep_pn)

    # 3. Main processing loop
    def get_header_cell(c, cmd):
        label = c.get("step_label") or f"Processing {c.get('header')}"
        if cmd and "ACRD" in cmd and "Calculate based on ACRD" in label: label = "Calculate Recommended Qty"
        return CellData(value=c["header"], stepLabel=label, agent=c.get("agent", "Data Mapping Agent"), tool=c.get("tool", "CrossRef_Engine"))

    header_row = [get_header_cell(c, command) for c in scenario["target_columns"]]
    res_preview, res_export = [header_row], [header_row]

    total_rows = len(source_rows) - 1
    for i in range(1, len(source_rows)):
        try:
            if i % 100 == 0 or i == 1: print(f"DEBUG: Processing row {i}/{total_rows} for scenario '{command}'")
            s_row = source_rows[i]
            p_row, e_row = [], []
            
            for col_idx, col in enumerate(scenario["target_columns"]):
                ctype = col["type"]
                info = {"dict": col.get("dict", {})}
                val, excel_formula = "", None
                if ctype == "source":
                    idx = get_column_index(source_headers, col.get("column", "").upper())
                    val = s_row[idx] if idx != -1 and idx < len(s_row) else ""
                elif ctype == "vlookup":
                    s_key_idx = col.get("search_key_col_idx", 0)
                    # Safe key retrieval
                    key_val = ""
                    if s_key_idx < len(p_row): key_val = p_row[s_key_idx].value
                    elif s_key_idx < len(s_row): key_val = s_row[s_key_idx]
                    
                    key = str(key_val).strip()
                    if key.endswith(".0"): key = key[:-2]
                    val = info.get("dict", {}).get(key, col.get("fallback_value", "N/A"))
                    suffix = col.get("formula_suffix")
                    if suffix and val != "N/A":
                        try: val = eval(f"{clean_num(val)}{suffix}")
                        except: pass
                    
                    l_file = col.get("lookup_file", "")
                    k_cell = f"{chr(65+s_key_idx)}{i+1}"
                    f_name = next((k for k in files_data.keys() if k.lower() == l_file.lower()), l_file)
                    
                    # Determine VLOOKUP range start column based on lookup_key_col_idx
                    l_k_idx = col.get("lookup_key_col_idx", 0)
                    l_start_char = chr(65 + l_k_idx)
                    
                    manual_range = col.get("manual_range")
                    if manual_range:
                        k_idx = manual_range.get("key_col_idx", 0)
                        v_idx = manual_range.get("val_col_idx", 1)
                        end_idx = manual_range.get("end_col_idx", v_idx)
                        k_char = get_column_letter(k_idx + 1)
                        v_char = get_column_letter(end_idx + 1)
                        # VLOOKUP index is relative to range start
                        lookup_idx = v_idx - k_idx + 1
                        excel_formula = f"VLOOKUP({k_cell}{separator} '{f_name}'!${k_char}$1:${v_char}$8000{separator} {lookup_idx}{separator} FALSE)"
                    else:
                        v_idx = col.get("target_col_idx", 1)
                        v_idx_rel = v_idx - l_k_idx + 1
                        f_suffix = suffix.replace(",", separator).replace(";", separator) if suffix else ""
                        excel_formula = f"VLOOKUP({k_cell}{separator} '{f_name}'!${l_start_char}$1:$Z$5000{separator} {v_idx_rel}{separator} FALSE){f_suffix}"
                    
                    # Wrap in IFERROR if fallback is 0 or numeric
                    f_val = col.get("fallback_value")
                    if f_val == 0 or f_val == "0" or (isinstance(f_val, (int, float)) and f_val == 0):
                        excel_formula = f"IFERROR({excel_formula}{separator} 0)"
                    
                    # Wrap ESS in INT()
                    if col.get("header") == "ESS":
                        excel_formula = f"INT({excel_formula})"
                    
                    if excel_formula:
                        excel_formula = excel_formula.replace(",", separator).replace(";", separator)
                elif ctype == "expression":
                    try:
                        eval_expr = col["formula_template"]
                        if "VLOOKUP" in eval_expr:
                            d_val = str(s_row[3]).strip() if len(s_row) > 3 else "0"
                            if d_val.endswith(".0"): d_val = d_val[:-2]
                            l_val = info.get("dict", {}).get(d_val, 0)
                            eval_expr = re.sub(r'VLOOKUP\s*\(\s*\{D\}\s*,[^)]+\)', str(l_val), eval_expr, flags=re.IGNORECASE)
                        for char in re.findall(r'\{([A-Z])\}', eval_expr):
                            idx = ord(char) - 65
                            val_ref = p_row[idx].value if idx < len(p_row) else 0
                            
                            # Detect if it's a numeric string or a real number
                            is_num = False
                            try:
                                if isinstance(val_ref, (int, float)): is_num = True
                                else:
                                    test_val = str(val_ref).replace(",", "").replace("$", "").strip()
                                    float(test_val)
                                    is_num = True
                            except: pass

                            if is_num:
                                eval_expr = eval_expr.replace(f'{{{char}}}', str(clean_num(val_ref)))
                            else:
                                # Treat as string for comparison
                                clean_str = str(val_ref).replace('"', '\\"')
                                eval_expr = eval_expr.replace(f'{{{char}}}', f'"{clean_str}"')
                        safety_if = 0
                        while "IF(" in eval_expr and safety_if < 20:
                            safety_if += 1
                            match = re.search(r'IF\(', eval_expr)
                            if not match: break
                            start_idx, bracket_count, current_pos = match.start(), 0, match.start() + 3
                            args, current_arg = [], ""
                            while current_pos < len(eval_expr):
                                char = eval_expr[current_pos]
                                if char == '(': bracket_count += 1
                                elif char == ')':
                                    if bracket_count == 0: args.append(current_arg.strip()); break
                                    bracket_count -= 1
                                elif char == ',' and bracket_count == 0:
                                    args.append(current_arg.strip()); current_arg = ""; current_pos += 1; continue
                                current_arg += char; current_pos += 1
                            if len(args) == 3:
                                cond, v_t, v_f = args
                                py_cond = cond.replace('=', '==').replace('>==', '>=').replace('<==', '<=').replace('!==', '!=')
                                python_if = f"({v_t} if {py_cond} else {v_f})"
                                eval_expr = eval_expr[:start_idx] + python_if + eval_expr[current_pos+1:]
                            else: break
                        eval_expr = eval_expr.replace("POWER", "pow").replace("ROUND", "round").replace("POISSON", "poisson_helper").replace("MAX", "max").replace("MIN", "min")
                        eval_context = {
                            "pow": pow, 
                            "round": round, 
                            "poisson_helper": poisson_helper, 
                            "max": max, 
                            "min": min, 
                            "AND": lambda *args: all(args[0] if len(args) == 1 and isinstance(args[0], (list, tuple)) else args),
                            "OR": lambda *args: any(args[0] if len(args) == 1 and isinstance(args[0], (list, tuple)) else args),
                            "TRUE": True, 
                            "FALSE": False
                        }
                        val = eval(eval_expr, eval_context)
                        
                        # Ensure excel_formula is assigned and fixed for O365
                        excel_formula = col.get("manual_formula") or col.get("formula_template")
                        if excel_formula:
                            excel_formula = excel_formula.replace("{i}", str(i+1))
                            excel_formula = excel_formula.replace(",", separator).replace(";", separator).replace("'", '"')
                            if separator == ",":
                                if "LET(" in excel_formula:
                                    excel_formula = excel_formula.replace("LET(", "_xlfn.LET(")
                                if "SEQUENCE(" in excel_formula:
                                    excel_formula = excel_formula.replace("SEQUENCE(", "_xlfn.SEQUENCE(")
                    except Exception as e:
                        import traceback
                        print(f"DEBUG: Eval failed for expression: {eval_expr}")
                        print(traceback.format_exc())
                        val = "Error"
                elif ctype == "poisson_spare":
                    # 1. Get Tolerance from Parameters!$F$1:$G$4 based on ESS
                    ess_val = clean_num(p_row[3].value) # Index 3 is ESS (D)
                    tolerance = 0
                    p_data = normalized_files.get('parameters')
                    if p_data:
                        for r in p_data:
                            # Search in col F (index 5) for ESS, get from col G (index 6)
                            if len(r) > 6 and clean_num(r[5]) == ess_val:
                                tolerance = clean_num(r[6])
                                break
                    
                    l_v = clean_num(p_row[11].value) # Index 11 is Expected Demand (L)
                    pl_v = clean_num(p_row[12].value) # Index 12 is Protection Level (M)
                    if pl_v > 1: pl_v /= 100.0
                    
                    target_v = pl_v - tolerance
                    qty, safety = 0, 0
                    while round(poisson_helper(qty, l_v, True), 2) < target_v and safety < 500:
                        qty += 1
                        safety += 1
                    val = qty
                    
                    # Ensure excel_formula is assigned and fixed for O365
                    excel_formula = col.get("manual_formula") or col.get("formula_template")
                    if excel_formula:
                        excel_formula = excel_formula.replace("{i}", str(i+1))
                        excel_formula = excel_formula.replace(",", separator).replace(";", separator).replace("'", '"')
                        if separator == ",":
                            excel_formula = excel_formula.replace("LET(", "_xlfn.LET(").replace("SEQUENCE(", "_xlfn.SEQUENCE(")
                    
                    if i <= 10:
                        print(f"DEBUG Row {i}: ESS={ess_val}, Tolerance={tolerance}, Target_V={target_v}, Qty={val}, Formula={excel_formula}")
                elif ctype in ["acrd_replacement_sum", "acrd_replacement_list"]:
                    s_key_idx = col.get("search_key_col_idx", 0)
                    pn_val = str(s_row[s_key_idx]).strip().upper() if s_key_idx < len(s_row) else ""
                    if pn_val.endswith(".0"): pn_val = pn_val[:-2]
                    reps = acrd_map.get(pn_val, [])
                    if ctype == "acrd_replacement_list": val = ", ".join(reps) if reps else ""
                    else:
                        q_orig, q_reps = pool_qty_map.get(pn_val, 0), sum(pool_qty_map.get(r_pn, 0) for r_pn in reps)
                        val = q_orig + q_reps
                
                elif ctype == "mock_file_list":
                    lines = info.get("lines", [])
                    # i is row index starting from 1 (headers at 0)
                    # We want line i-1 for row i
                    if (i-1) < len(lines):
                        val = lines[i-1]
                        try:
                            # Convert to number if possible for better Excel handling
                            if '.' in val: val = float(val)
                            else: val = int(val)
                        except: pass
                
                display_val = val
                if col.get("format") == "currency":
                    try: display_val = f"${float(val):,.2f}"
                    except: pass
                p_row.append(CellData(value=display_val))
                
                e_row.append(CellData(value=val, formula=excel_formula))
                if i == 1 and excel_formula:
                    print(f"DEBUG SAMPLE FORMULA (Col {col.get('header')}): {excel_formula}")
            
            res_preview.append(p_row); res_export.append(e_row)
            
            # Debug first 10 rows
            if i <= 10:
                try:
                    exp_demand = p_row[11].value
                    qty_val = p_row[15].value
                    print(f"DEBUG Row {i}: Expected Demand={exp_demand}, Qty={qty_val}")
                except: pass
        except Exception as row_err:
            import traceback
            print(f"ERROR: Failed to process row {i}: {row_err}")
            print(traceback.format_exc())
            continue

    # 4. Final summary and ESS metrics
    col_map = {h.value: i for i, h in enumerate(res_preview[0])}
    impact_summary = {"partsProcessed": len(res_preview) - 1, "status": "success", "sparesRecommended": 0, "essMetrics": {}, "plan": []}
    for c in scenario["target_columns"]:
        label = c.get("step_label") or f"Processing {c.get('header')}"
        if command and "ACRD" in command and "Calculate based on ACRD" in label: label = "Calculate Recommended Qty"
        impact_summary["plan"].append({"phase": c.get("phase", "Modeling"), "goal": label, "agent": c.get("agent", "Data Mapping Agent"), "tool": c.get("tool", "CrossRef_Engine"), "status": "completed"})

    if command in ["Provisioning", "VLOOKUP", "Pool Buy Scenarios"]:
        ess_sums = {"ess1_ar": 0.0, "ess1_ac": 0.0, "ess2_ar": 0.0, "ess2_ac": 0.0, "ess3_ar": 0.0, "ess3_ac": 0.0}
        qty_header = "Qty" if command != "Pool Buy Scenarios" else "Pool Buy Qty"
        for row in res_preview[1:]:
            for k in ["ESS1", "ESS2", "ESS3"]:
                ess_sums[f"{k.lower()}_ar"] += clean_num(row[col_map.get(f"{k} Arising TRT")].value) if f"{k} Arising TRT" in col_map else 0
                ess_sums[f"{k.lower()}_ac"] += clean_num(row[col_map.get(f"{k} Actual TRT Demand")].value) if f"{k} Actual TRT Demand" in col_map else 0
            impact_summary["sparesRecommended"] += clean_num(row[col_map.get(qty_header)].value) if qty_header in col_map else 0
        
        if command == "Pool Buy Scenarios":
            total_buy = int(impact_summary["sparesRecommended"])
            total_value = sum(clean_num(row[col_map.get("Pool Buy (USD)")].value) for row in res_preview[1:] if "Pool Buy (USD)" in col_map)
            impact_summary["totalBuyValue"] = total_value
            impact_summary["businessRationale"] = f"Based on the combined analysis of Current MBH Fleet, Standalone IP1, and Manual Top-ups, the total recommended purchase is {total_buy} units with an estimated value of ${total_value:,.2f}. [Verified]"
        
        actual_params, p_d = {"ess1": 98, "ess2": 95, "ess3": 92}, normalized_files.get('parameters')
        if p_d:
            try:
                actual_params["ess1"] = int(round(clean_num(p_d[8][2]) * 100))
                actual_params["ess2"] = int(round(clean_num(p_d[9][2]) * 100))
                actual_params["ess3"] = int(round(clean_num(p_d[10][2]) * 100))
            except: pass
        targets = {"ess1": {"val": 98.0, "cell": "C9", "real": actual_params["ess1"]}, "ess2": {"val": 95.0, "cell": "C10", "real": actual_params["ess2"]}, "ess3": {"val": 92.0, "cell": "C11", "real": actual_params["ess3"]}}
        for key, t in targets.items():
            ar, ac = ess_sums[f"{key}_ar"], ess_sums[f"{key}_ac"]
            rate = (ac / ar * 100) if ar > 0 else 100.0
            status = "Passed" if rate >= t["val"] else "Failed"
            recom = f"Recommend update Protection Level in cell {t['cell']} of Parameter sheet from {t['real']}% to {t['real']+1}%." if status == "Failed" else ""
            impact_summary["essMetrics"][key] = {"actual": rate, "desired": t["val"], "status": status, "recommendation": recom, "targetCell": t["cell"], "suggestedValue": (t["real"] + 1) / 100.0 if status == "Failed" else None}
    elif command and "ACRD" in command:
        h = "Adjusted Total Avail Pool based on Pool Std"
        impact_summary["sparesRecommended"] = int(sum(clean_num(row[col_map.get(h)].value) for row in res_preview[1:] if h in col_map))

    # Build Final Workbooks with ALL sheets
    result_sheet_name = command if command else "Provisioning"
    
    # Preview WB needs Source + Result for the UI typing effect to work correctly
    source_cells = [[CellData(value=c) for c in r] for r in source_rows]
    preview_sheets = [
        SheetData(name="PartNumbers", data=source_cells),
        SheetData(name="VLOOKUP_Result", data=res_preview)
    ]
    preview_wb = WorkbookData(sheets=preview_sheets, activeSheetIndex=1)

    export_sheets = [SheetData(name=result_sheet_name, data=res_export)]
    for sheet_name, sheet_data in files_data.items():
        if sheet_name.lower() != result_sheet_name.lower():
            cells = [[CellData(value=c) for c in r] for r in sheet_data]
            export_sheets.append(SheetData(name=sheet_name, data=cells))
            
    if len(res_preview) > 1:
        print(f"DEBUG: Returning {len(res_preview)} rows. First data cell (1,0): {res_preview[1][0].value}")
        
        # Calculate and print sums for verification
        headers = [str(h.value) for h in res_preview[0]]
        col_sums = {h: 0.0 for h in headers}
        for row in res_preview[1:]:
            for idx, cell in enumerate(row):
                val = clean_num(cell.value)
                if isinstance(val, (int, float)):
                    col_sums[headers[idx]] += val
        
        print(f"DEBUG: === COLUMN SUMS FOR SCENARIO '{command}' ===")
        for h, s in col_sums.items():
            if s != 0: # Only print columns that actually have numeric data
                print(f"  - {h}: {round(s, 2)}")
        print("DEBUG: ===============================================")
    
    export_wb = WorkbookData(sheets=export_sheets, activeSheetIndex=0)
    return preview_wb, export_wb, impact_summary

def generate_fallback_workbooks(command: str, prompt: str):
    wb = WorkbookData(sheets=[SheetData(name="Result", data=[[CellData(value="N/A")]])], activeSheetIndex=0)
    return wb, wb, {"status": "fallback"}
